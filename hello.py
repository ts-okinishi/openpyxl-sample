# coding:utf-8

from openpyxl import Workbook

# 新規ワークブック作成
wb = Workbook()

# アクティブワークシート取得
ws = wb.active

# 値を入力する
ws['A1'] = 42
ws.cell(row=1, column=2, value='こんにちは') # B1

for c in ws['A2:C3']:
    c = 'まとめて'

# まとめて最終行に入力する
ws.append([1, 2, 3])

# 日付
import datetime
ws['A2'] = datetime.datetime.now()

# 新規ワークシートを作成する
ws1 = wb.create_sheet('NewSheet', 1)

ws2 = wb.create_sheet('NewSheet2', 0)
ws2.title = 'NewSheet2!!' # ValueError: Worksheet titles must be unicode
ws2.sheet_properties.tabColor = '1072BA'
ws2['A1'] = 'A1'

ws3 = wb.copy_worksheet(ws2)

ws4 = wb['NewSheet']
ws4['A1'] = 'これは左から3番目にあるシート'

# ファイル保存
wb.save("sample.xlsx")
