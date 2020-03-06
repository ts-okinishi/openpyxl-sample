# coding:utf-8

from openpyxl import load_workbook

# ワークブック読み込み
wb = load_workbook('sample.xlsx')

# アクティブワークシート取得
ws = wb.active

# 値を入力する
ws['A1'] = '値を変更！'

# 別ファイルに保存
wb.save("sample1.xlsx")
