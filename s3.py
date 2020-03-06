# coding:utf-8

from openpyxl import load_workbook
from tempfile import NamedTemporaryFile
import boto3 # `pip install boto3` or `pip3 install boto3`
import botocore

s3 = boto3.resource('s3')

# S3のファイルをダウンロードする
def download_file(bucket, key):
    try:
        file = NamedTemporaryFile(suffix = '.xlsx', delete=False)
        s3.Bucket(bucket).download_file(key, file.name)
        return file.name
    except botocore.exceptions.ClientError as e:
        if e.response['Error']['Code'] == "404":
            return None
        else:
            raise
    else:
        raise

# S3へファイルをアップロードする
def upload_workbook(workbook, bucket, key):
    with NamedTemporaryFile() as tmp:
        workbook.save(tmp.name)
        tmp.seek(0)
        s3.meta.client.upload_file(tmp.name, bucket, key)

# 注意: バケットを先に作っておく

# S3からワークブックをダウンロード
file = download_file("test-bucket", "sample2.xlsx")
wb = load_workbook(file)

# ワークブックをS3へアップロード
upload_workbook(wb, "test-bucket", "サンプル2.xlsx")
